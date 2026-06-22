import argparse
import csv
import os
import sys
from pathlib import Path

from dotenv import load_dotenv
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from classify import classify
from ingest import extract_text, get_listing_id, load_listings
from schema import requires_human_review

load_dotenv()

CATEGORY_FILL = {
    "Nursery": PatternFill("solid", fgColor="FCE4EC"),    # light pink
    "SEN School": PatternFill("solid", fgColor="BDD7EE"), # blue
    "Food Store": PatternFill("solid", fgColor="FCE4D6"), # orange
    "None": PatternFill(fill_type=None),
}
CONFIDENCE_FILL = {
    "High": PatternFill("solid", fgColor="C6EFCE"),   # green
    "Medium": PatternFill("solid", fgColor="FFEB9C"), # yellow
    "Low": PatternFill("solid", fgColor="FFC7CE"),    # red
}


def write_xlsx(pairs: list, path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["listing_id", "category", "confidence", "reasoning", "requires_human_review"])

    for _, result in pairs:
        review = requires_human_review(result.confidence)
        ws.append([
            result.listing_id,
            result.category,
            result.confidence,
            result.reasoning,
            review,
        ])
        r = ws.max_row
        ws.cell(r, 2).fill = CATEGORY_FILL.get(result.category, PatternFill(fill_type=None))
        ws.cell(r, 3).fill = CONFIDENCE_FILL.get(result.confidence, PatternFill(fill_type=None))

    wb.save(path)


def write_csv(pairs: list, path: Path) -> None:
    if not pairs:
        return
    original_keys = list(pairs[0][0].keys())
    fieldnames = original_keys + ["classified_category", "confidence", "reasoning", "requires_human_review"]
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row, result in pairs:
            writer.writerow({
                **row,
                "classified_category": result.category,
                "confidence": result.confidence,
                "reasoning": result.reasoning,
                "requires_human_review": requires_human_review(result.confidence),
            })


def main() -> None:
    parser = argparse.ArgumentParser(description="Classify unseen property listings.")
    parser.add_argument("input", type=Path, help="Path to input CSV")
    parser.add_argument("--output", type=Path, default=None, help="Path to output Excel file")
    args = parser.parse_args()

    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        print("Error: OPENAI_API_KEY not set.", file=sys.stderr)
        sys.exit(1)

    output = args.output or Path("results") / f"{args.input.stem}_classified.xlsx"
    output.parent.mkdir(exist_ok=True)

    client = OpenAI(api_key=api_key)
    rows = load_listings(args.input)
    print(f"Loaded {len(rows)} listings from {args.input}\n")

    pairs = []
    for i, row in enumerate(rows):
        listing_id = get_listing_id(row, i)
        text = extract_text(row)
        print(f"[{i + 1}/{len(rows)}] {listing_id} ...", end=" ", flush=True)
        result = classify(client, listing_id, text)
        print(f"{result.category} ({result.confidence})")
        pairs.append((row, result))

    csv_output = output.with_suffix(".csv")
    write_xlsx(pairs, output)
    write_csv(pairs, csv_output)
    print(f"\nResults written to {output} and {csv_output}")


if __name__ == "__main__":
    main()
