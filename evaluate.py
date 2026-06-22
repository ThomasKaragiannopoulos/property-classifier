import csv
import os
import sys
from pathlib import Path

from dotenv import load_dotenv
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from classify import classify
from ingest import extract_text, get_listing_id, load_listings

load_dotenv()

INPUT = Path("data/listings.csv")
BASELINE = Path("data/manual_baseline.csv")
RESULTS_DIR = Path("results")
OUTPUT_CSV = RESULTS_DIR / "predictions.csv"
OUTPUT_XLSX = RESULTS_DIR / "predictions.xlsx"

GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")


def load_baseline(path: Path) -> dict:
    with open(path, newline="", encoding="utf-8") as f:
        return {row["listing_id"]: row for row in csv.DictReader(f)}



def write_csv(pairs: list, original_fields: list, path: Path) -> None:
    fieldnames = original_fields + ["classified_category", "confidence", "reasoning"]
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row, result in pairs:
            out = dict(row)
            out["classified_category"] = result.category
            out["confidence"] = result.confidence
            out["reasoning"] = result.reasoning
            writer.writerow(out)


def write_xlsx(pairs: list, baseline: dict, path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append([
        "listing_id", "category", "confidence",
        "reasoning", "manual_categorisation",
    ])

    for _, result in pairs:
        b = baseline.get(result.listing_id, {})

        ws.append([
            result.listing_id,
            result.category,
            result.confidence,
            result.reasoning,
            b.get("category", ""),
        ])

        r = ws.max_row

        cat_match = result.category == b.get("category", "")
        ws.cell(r, 2).fill = GREEN if cat_match else RED

        conf_match = result.confidence == b.get("confidence", "")
        ws.cell(r, 3).fill = GREEN if conf_match else YELLOW


    wb.save(path)


def main() -> None:
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        print("Error: OPENAI_API_KEY not set.", file=sys.stderr)
        sys.exit(1)

    client = OpenAI(api_key=api_key)
    rows = load_listings(INPUT)
    baseline = load_baseline(BASELINE)
    RESULTS_DIR.mkdir(exist_ok=True)
    print(f"Loaded {len(rows)} listings.\n")

    pairs = []
    for i, row in enumerate(rows):
        listing_id = get_listing_id(row, i)
        listing_text = extract_text(row)

        print(f"[{i + 1}/{len(rows)}] {listing_id} ...", end=" ", flush=True)
        result = classify(client, listing_id, listing_text)
        print(f"{result.category} ({result.confidence})")

        pairs.append((row, result))

    write_csv(pairs, list(rows[0].keys()), OUTPUT_CSV)
    write_xlsx(pairs, baseline, OUTPUT_XLSX)
    print(f"\nResults written to {OUTPUT_CSV} and {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
